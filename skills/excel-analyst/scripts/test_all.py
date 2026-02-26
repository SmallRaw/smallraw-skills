#!/usr/bin/env python3
"""
一键测试 excel_tool.py 全部功能。
使用前确保已安装 openpyxl：pip install openpyxl

运行：python scripts/test_all.py
"""

import os
import sys
import json
import subprocess

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.join(SCRIPT_DIR, "..")
TEST_DIR = os.path.join(SKILL_DIR, "test_output")
TOOL = os.path.join(SCRIPT_DIR, "excel_tool.py")
PYTHON = sys.executable

os.makedirs(TEST_DIR, exist_ok=True)


def run(cmd, label):
    print(f"\n{'='*60}")
    print(f"测试: {label}")
    print(f"命令: {' '.join(cmd)}")
    print(f"{'='*60}")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.stdout:
        print(result.stdout)
    if result.stderr:
        print(f"[STDERR] {result.stderr}")
    print(f"[退出码] {result.returncode}")
    return result.returncode == 0


def step1_generate_test_excel():
    """生成一个模拟乱报表的 Excel 文件"""
    print("\n" + "=" * 60)
    print("步骤 1: 生成测试 Excel 文件")
    print("=" * 60)

    try:
        import openpyxl
    except ImportError:
        print("错误：需要 openpyxl。运行 pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 乱报表（表头不在第一行）----
    ws1 = wb.active
    ws1.title = "销售月报"

    # 前3行是标题和元信息
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "2024年Q3季度电脑销售报表"
    ws1["A2"] = "制表日期"
    ws1["B2"] = "2024-10-15"
    ws1["A3"] = ""  # 空行

    # 第4行是真正的表头（0-based: row 3）
    headers = ["序号", "区域", "产品型号", "类别", "销量", "单价", "营收"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=4, column=c, value=h)

    # 数据（故意加入脏数据）
    data = [
        [1, "华东", "ThinkPad X1", "笔记本", 320, 8999, 2879680],
        [2, "华东", "Dell G15", "笔记本", 280, 6999, 1959720],
        [3, "华南区", "ThinkPad X1", "笔记本", 250, 8999, 2249750],  # "华南区"不统一
        [4, "华南", "MacBook Pro", "笔记本", 420, 12999, 5459580],
        [5, "华东区", "Dell G15", "笔记本", 180, 6999, 1259820],    # "华东区"不统一
        [6, "华北", "联想拯救者", "笔记本", 350, 7999, 2799650],
        [7, "华北", "HP战66", "台式机", 200, 4999, 999800],
        [8, "华南", "ThinkPad X1", "笔记本", None, 8999, None],      # 空值
        [9, "华东", "MacBook\nAir", "笔记本", 150, 9999, 1499850],    # 型号中间有换行
        [10, " 华北 ", "Dell\r\nG15", "笔记本", 120, 6999, 839880], # 带空格 + \r\n
        [11, "华东", "ThinkPad X1", "笔记本", 320, 8999, 2879680],  # 重复行
        [12, "华南地区", "HP\u200b战66", "台式机", 90, 4999, 449910],  # 零宽空格
        # 更多脏字符测试
        [16, "\xa0华东\xa0", "Surface\tPro", "笔记本", 100, 9999, 999900],  # NBSP + 制表符
        [17, "\u3000西南\u3000", "ThinkPad\n X1", "笔记本", 80, 8999, 719920],  # 全角空格 + 换行带空格
        [18, "华北", "\ufeffDell G15", "笔记本", 60, 6999, 419940],  # BOM
        [13, "西南", "联想拯救者", "笔记本", 280, 7999, 2239720],
        [14, "西南", "MacBook Pro", "笔记本", 190, 12999, 2469810],
        [15, "华北", "组装台式机", "台式机", 500, 3500, 1750000],
    ]

    for r_idx, row in enumerate(data, 5):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)

    # ---- Sheet 2: 简单表 ----
    ws2 = wb.create_sheet("库存")
    ws2.append(["型号", "库存数量", "仓库"])
    ws2.append(["ThinkPad X1", 500, "上海"])
    ws2.append(["Dell G15", 300, "深圳"])
    ws2.append(["MacBook Pro", 200, "北京"])

    test_file = os.path.join(TEST_DIR, "测试报表.xlsx")
    wb.save(test_file)
    print(f"已生成: {test_file}")
    return test_file


def step2_test_scout(test_file):
    """测试侦察功能"""
    return run(
        [PYTHON, TOOL, "scout", test_file, "-n", "6"],
        "scout - 侦察前6行"
    )


def step3_test_auto_no_config(test_file):
    """测试自动模式（无配置，应触发侦察）"""
    # 确保无配置
    cfg = os.path.splitext(test_file)[0] + ".excel-config.json"
    if os.path.exists(cfg):
        os.remove(cfg)

    return run(
        [PYTHON, TOOL, "auto", test_file, "preview"],
        "auto preview - 无配置时自动侦察"
    )


def step4_generate_config(test_file):
    """手动生成结构配置"""
    print(f"\n{'='*60}")
    print("步骤 4: 生成结构配置")
    print("=" * 60)

    config = {
        "sheets": {
            "销售月报": {
                "header_row": 3,
                "data_start_row": 4,
                "columns": {
                    "产品型号": "型号"
                },
                "skip_cols": [0],
                "notes": "前3行是标题，第0列序号跳过"
            }
        }
    }

    cfg_path = os.path.splitext(test_file)[0] + ".excel-config.json"
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"已生成: {cfg_path}")
    print(json.dumps(config, ensure_ascii=False, indent=2))
    return True


def step5_test_auto_with_config(test_file):
    """测试有配置的各种查询"""
    ok = True
    ok &= run(
        [PYTHON, TOOL, "auto", test_file, "headers"],
        "auto headers - 查看表头"
    )
    ok &= run(
        [PYTHON, TOOL, "auto", test_file, "preview", "-n", "5"],
        "auto preview - 预览前5行"
    )
    ok &= run(
        [PYTHON, TOOL, "auto", test_file, "query",
         "--where-col", "类别", "--where-op", "==", "--where-val", "笔记本",
         "-s", "desc:销量", "-t", "5"],
        "auto query - 筛选笔记本按销量降序取前5"
    )
    ok &= run(
        [PYTHON, TOOL, "auto", test_file, "query",
         "--where-col", "销量", "--where-op", ">", "--where-val", "300",
         "-c", "区域,型号,销量"],
        "auto query - 销量>300，只看区域/型号/销量"
    )
    return ok


def step6_test_clean(test_file):
    """测试数据清洗"""
    print(f"\n{'='*60}")
    print("步骤 6: 生成清洗规则并测试")
    print("=" * 60)

    rules = {
        "steps": [
            {"action": "trim"},
            {
                "action": "replace",
                "column": "区域",
                "mapping": {
                    "华东区": "华东",
                    "华南区": "华南",
                    "华南地区": "华南"
                }
            },
            {"action": "fill_empty", "column": "销量", "value": 0},
            {"action": "fill_empty", "column": "营收", "value": 0},
            {"action": "dedup", "columns": ["区域", "型号", "销量"]},
            {
                "action": "filter",
                "conditions": [
                    {"column": "类别", "op": "==", "value": "笔记本"}
                ],
                "logic": "and"
            },
            {
                "action": "aggregate",
                "group_by": ["区域"],
                "metrics": {
                    "销量": "sum",
                    "营收": "sum"
                }
            },
            {"action": "sort", "column": "销量", "desc": True}
        ]
    }

    rules_path = os.path.join(TEST_DIR, "测试清洗规则.json")
    with open(rules_path, "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)
    print(f"已生成规则: {rules_path}")

    ok = True

    # 预览
    ok &= run(
        [PYTHON, TOOL, "clean", test_file, rules_path, "--preview"],
        "clean --preview - 预览清洗结果"
    )

    # 导出 CSV
    csv_out = os.path.join(TEST_DIR, "清洗结果.csv")
    ok &= run(
        [PYTHON, TOOL, "clean", test_file, rules_path, "-o", csv_out],
        "clean -o csv - 导出CSV"
    )

    # 导出 JSON
    json_out = os.path.join(TEST_DIR, "清洗结果.json")
    ok &= run(
        [PYTHON, TOOL, "clean", test_file, rules_path, "-o", json_out],
        "clean -o json - 导出JSON"
    )

    # 导出 Excel
    xlsx_out = os.path.join(TEST_DIR, "清洗结果.xlsx")
    ok &= run(
        [PYTHON, TOOL, "clean", test_file, rules_path, "-o", xlsx_out],
        "clean -o xlsx - 导出Excel"
    )

    # 验证导出文件
    print(f"\n{'='*60}")
    print("验证导出文件")
    print("=" * 60)
    for f in [csv_out, json_out, xlsx_out]:
        exists = os.path.exists(f)
        size = os.path.getsize(f) if exists else 0
        status = f"存在 ({size} bytes)" if exists else "缺失"
        print(f"  {os.path.basename(f)}: {status}")

    return ok


def main():
    print("=" * 60)
    print("  Excel Tool 全功能测试")
    print("=" * 60)

    test_file = step1_generate_test_excel()
    results = {}

    results["scout"] = step2_test_scout(test_file)
    results["auto_no_config"] = step3_test_auto_no_config(test_file)
    results["gen_config"] = step4_generate_config(test_file)
    results["auto_headers"] = step5_test_auto_with_config(test_file)
    results["clean"] = step6_test_clean(test_file)

    print(f"\n\n{'='*60}")
    print("  测试汇总")
    print("=" * 60)
    for name, ok in results.items():
        print(f"  {name}: {'PASS' if ok else 'FAIL'}")

    all_pass = all(results.values())
    print(f"\n  总结: {'全部通过' if all_pass else '有失败项'}")
    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())
