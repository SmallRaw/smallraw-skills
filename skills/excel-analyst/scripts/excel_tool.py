#!/usr/bin/env python3
"""
Excel 报表工具 - 处理复杂/乱序报表
读取层：pywin32（Windows）/ openpyxl（跨平台） - 处理原始结构
处理层：pandas - 查询、清洗、聚合
"""

import os
import sys
import json
import argparse
import re

import pandas as pd

# ==================== 引擎检测（仅用于读取原始结构）====================

READ_ENGINE = None

try:
    import win32com.client
    READ_ENGINE = "pywin32"
except ImportError:
    pass

if READ_ENGINE is None:
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        READ_ENGINE = "openpyxl"
    except ImportError:
        print("错误：需要安装 pywin32（Windows）或 openpyxl（跨平台）")
        sys.exit(1)

# ==================== 路径与配置 ====================

TOOL_PATH = os.path.abspath(__file__)

CONFIG_TEMPLATE = """{
  "sheets": {
    "Sheet名称": {
      "header_row": 0,
      "data_start_row": 1,
      "columns": {"原始列名": "标准化名称"},
      "skip_cols": [],
      "notes": "一句话说明"
    }
  }
}"""

STEPS_TEMPLATE = """{
  "steps": [
    {"action": "trim"},
    {"action": "dedup"}
  ]
}"""


def get_config_path(excel_path):
    abs_path = os.path.abspath(excel_path)
    name = os.path.splitext(abs_path)[0]
    return f"{name}.excel-config.json"


def get_steps_pattern(excel_path):
    """返回匹配该 Excel 所有 steps 文件的 glob 模式"""
    abs_path = os.path.abspath(excel_path)
    name = os.path.splitext(abs_path)[0]
    return f"{name}*.excel-steps.json"


def get_steps_prefix(excel_path):
    """返回该 Excel 对应 steps 文件的前缀（不含后缀），用于生成新文件名"""
    abs_path = os.path.abspath(excel_path)
    return os.path.splitext(abs_path)[0]


def discover_steps_files(excel_path):
    """发现该 Excel 关联的所有 .excel-steps.json 文件"""
    import glob
    pattern = get_steps_pattern(excel_path)
    return sorted(glob.glob(pattern))


def _load_config(config_path):
    """加载配置文件"""
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def _get_excel_sheet_names(file_path):
    """获取 Excel 文件中所有 Sheet 名称"""
    if READ_ENGINE == "pywin32":
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(file_path))
            names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
            wb.Close(False)
            return names
        finally:
            excel.Quit()
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names


# ==================== 文本统一清洗 ====================

# 零宽字符（不可见但影响匹配）→ 直接删除
_RE_ZERO_WIDTH = re.compile(r"[\u200b\u200c\u200d\u200e\u200f\u2060\ufeff]")
# 各种换行、制表、控制字符、特殊空白 → 替换为普通空格
_RE_CTRL_SPACE = re.compile(r"[\r\n\t\x0b\x0c\xa0\u3000]+")
# 连续空白 → 单空格
_RE_MULTI_SPACE = re.compile(r" {2,}")


def _clean_text(val):
    """统一清洗单个文本值：零宽字符→删除，控制字符/特殊空白→空格，连续空白→单空格，首尾去空"""
    if val is None:
        return None
    s = str(val)
    s = _RE_ZERO_WIDTH.sub("", s)
    s = _RE_CTRL_SPACE.sub(" ", s)
    s = _RE_MULTI_SPACE.sub(" ", s)
    return s.strip()


# ==================== 侦察：openpyxl / pywin32（需要看原始单元格）====================

def scout_pywin32(file_path, rows):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    result = {}
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        for i in range(1, wb.Sheets.Count + 1):
            ws = wb.Sheets(i)
            used = ws.UsedRange
            if used is None:
                continue
            total_rows = used.Rows.Count
            total_cols = used.Columns.Count
            start_row = used.Row
            start_col = used.Column
            lines = []
            for r in range(start_row, min(start_row + rows, start_row + total_rows)):
                cells = []
                for c in range(start_col, start_col + total_cols):
                    cell = ws.Cells(r, c)
                    val = _clean_text(cell.Text)
                    if not val:
                        val = "[空]"
                    if cell.MergeCells:
                        val = f"{val}[合并]"
                    cells.append(val)
                lines.append(cells)
            result[ws.Name] = {
                "total_rows": total_rows, "total_cols": total_cols,
                "start_row": start_row, "start_col": start_col,
                "preview": lines
            }
        wb.Close(False)
    finally:
        excel.Quit()
    return result


def scout_openpyxl(file_path, rows):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {}
    for ws in wb:
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        min_row = ws.min_row or 1
        min_col = ws.min_column or 1
        lines = []
        for row in ws.iter_rows(min_row=min_row,
                                max_row=min(min_row + rows - 1, total_rows),
                                values_only=False):
            cells = []
            for cell in row:
                val = _clean_text(cell.value)
                if not val:
                    val = "[空]"
                if isinstance(cell, MergedCell):
                    val = f"{val}[合并]"
                cells.append(val)
            lines.append(cells)
        result[ws.title] = {
            "total_rows": total_rows, "total_cols": total_cols,
            "start_row": min_row, "start_col": min_col,
            "preview": lines
        }
    wb.close()
    return result


def do_scout(file_path, rows=8, sheet=None):
    result = scout_pywin32(file_path, rows) if READ_ENGINE == "pywin32" else scout_openpyxl(file_path, rows)
    found = False
    for sheet_name, info in result.items():
        if sheet and sheet_name != sheet:
            continue
        found = True
        print(f"\n=== {sheet_name} "
              f"({info['total_rows']}行 × {info['total_cols']}列, "
              f"起始:R{info['start_row']}C{info['start_col']}) ===")
        for idx, row in enumerate(info["preview"]):
            display_row = info['start_row'] + idx
            config_row = display_row - 1
            print(f"  行{display_row}(config:{config_row}): {' | '.join(str(v) for v in row)}")
    if sheet and not found:
        available = list(result.keys())
        print(f"\n[错误] Sheet '{sheet}' 不存在")
        print(f"[可用 Sheet] {', '.join(available)}")
    elif found:
        print(f"\n[行号说明] 括号中 config:N 的值可直接用于配置文件的 header_row / data_start_row")
    return result


def _guess_config(scout_data):
    """分析侦察结果，为每个 Sheet 推荐配置（表头行、数据起始行、是否跳过序号列）"""
    config = {"sheets": {}}

    for sheet_name, info in scout_data.items():
        preview = info["preview"]
        start_row = info["start_row"]  # 1-based

        if not preview:
            config["sheets"][sheet_name] = {
                "header_row": 0, "data_start_row": 1,
                "columns": {}, "skip_cols": [],
                "notes": "空Sheet，请手动确认"
            }
            continue

        # 对每一行打分，选最可能的表头行
        best_idx = 0
        best_score = -1

        for idx, row in enumerate(preview):
            # 有合并单元格的行是标题，不是表头
            if any("[合并]" in str(c) for c in row):
                continue
            # 全空行不是表头
            non_empty = sum(1 for c in row if str(c) != "[空]")
            if non_empty == 0:
                continue
            score = non_empty / len(row) if row else 0
            if score > best_score:
                best_score = score
                best_idx = idx

        config_header = start_row + best_idx - 1  # 转 0-based
        config_data = config_header + 1

        # 检测第一列是否为序号列
        skip_cols = []
        header_row_cells = preview[best_idx]
        data_rows = preview[best_idx + 1:] if best_idx + 1 < len(preview) else []

        if header_row_cells and data_rows:
            first_header = str(header_row_cells[0]).replace("[空]", "").strip()
            serial_hints = ["序号", "编号", "No", "no", "NO", "#", "序", "行号", "ID", "id"]
            is_serial = any(h in first_header for h in serial_hints)

            if not is_serial and len(data_rows) >= 2:
                nums = []
                for row in data_rows:
                    val = str(row[0]).replace("[空]", "").strip()
                    try:
                        nums.append(int(float(val)))
                    except (ValueError, TypeError):
                        break
                is_serial = (
                    len(nums) >= 2
                    and all(nums[i] == nums[i - 1] + 1 for i in range(1, len(nums)))
                )

            if is_serial:
                skip_cols = [0]

        # 生成备注
        notes = []
        if best_idx > 0:
            notes.append(f"前{best_idx}行是标题/空行")
        if skip_cols:
            notes.append(f"第0列'{first_header}'为序号已跳过")

        config["sheets"][sheet_name] = {
            "header_row": config_header,
            "data_start_row": config_data,
            "columns": {},
            "skip_cols": skip_cols,
            "notes": "，".join(notes) if notes else "标准格式"
        }

    return config


# ==================== 读取为 DataFrame：openpyxl / pywin32 读原始数据 → pandas ====================

def _normalize_strings(df):
    """清理字符串列：零宽字符、控制字符、特殊空白统一处理"""
    str_cols = df.select_dtypes(include=["object", "str"]).columns
    for col in str_cols:
        df[col] = df[col].map(lambda v: _clean_text(v) if pd.notna(v) else v)
        df[col] = df[col].replace(["None", "nan", ""], pd.NA)
    return df


def read_to_dataframe(file_path, sheet_name, sheet_cfg):
    """读取 Excel 指定 Sheet 并返回 pandas DataFrame（自动清理脏字符）"""
    if READ_ENGINE == "pywin32":
        df = _read_pywin32_to_df(file_path, sheet_name, sheet_cfg)
    else:
        df = _read_openpyxl_to_df(file_path, sheet_name, sheet_cfg)
    return _normalize_strings(df)


def _read_pywin32_to_df(file_path, sheet_name, cfg):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        ws = wb.Sheets(sheet_name)
        header_row = cfg["header_row"] + 1
        data_start = cfg["data_start_row"] + 1
        used = ws.UsedRange
        end_row = used.Row + used.Rows.Count - 1
        end_col = used.Column + used.Columns.Count - 1
        skip_cols = set(cfg.get("skip_cols", []))
        col_map = cfg.get("columns", {})

        headers = []
        col_indices = []
        for c in range(1, end_col + 1):
            if (c - 1) in skip_cols:
                continue
            raw = _clean_text(ws.Cells(header_row, c).Text) or ""
            headers.append(col_map.get(raw, raw))
            col_indices.append(c)

        data = []
        for r in range(data_start, end_row + 1):
            row = [ws.Cells(r, c).Value for c in col_indices]
            if any(v is not None and str(v).strip() for v in row):
                data.append(row)

        wb.Close(False)
        return pd.DataFrame(data, columns=headers)
    finally:
        excel.Quit()


def _read_openpyxl_to_df(file_path, sheet_name, cfg):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header_row = cfg["header_row"] + 1
        data_start = cfg["data_start_row"] + 1
        skip_cols = set(cfg.get("skip_cols", []))
        col_map = cfg.get("columns", {})

        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
        headers = []
        col_indices = []
        for idx, cell in enumerate(header_cells):
            if idx in skip_cols:
                continue
            raw = _clean_text(cell.value) or ""
            headers.append(col_map.get(raw, raw))
            col_indices.append(idx)

        data = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            filtered = [row[i] for i in col_indices]
            if any(v is not None and str(v).strip() for v in filtered):
                data.append(filtered)

        return pd.DataFrame(data, columns=headers)
    finally:
        wb.close()


# ==================== auto 命令：pandas 查询 ====================

def do_auto(file_path, action="preview", sheet=None, **kwargs):
    config_path = get_config_path(file_path)

    if not os.path.exists(config_path):
        print(f"[引擎] {READ_ENGINE}")
        print(f"[状态] 首次分析该文件，需要先确定表格结构\n")
        scout_data = do_scout(file_path, sheet=sheet)

        # 自动推荐配置
        target_data = {k: v for k, v in scout_data.items() if k == sheet} if sheet else scout_data
        recommended = _guess_config(target_data)
        config_json = json.dumps(recommended, ensure_ascii=False, indent=2)

        abs_file = os.path.abspath(file_path)
        print(f"\n{'='*60}")
        print(f"[推荐配置] 工具已自动分析，请确认后保存")
        print(f"用 Write 工具将以下 JSON 保存到：{config_path}\n")
        print(config_json)
        first_sheet = list(recommended["sheets"].keys())[0] if recommended["sheets"] else ""
        sheet_opt = f' --sheet "{first_sheet}"' if first_sheet else ""
        print(f"\n[下一步] 保存配置后执行：")
        print(f"  python {TOOL_PATH} auto {abs_file} headers{sheet_opt}")
        return

    cfg = _load_config(config_path)
    sheets = cfg["sheets"]

    # 确定要操作的 sheet
    if sheet:
        if sheet not in sheets:
            excel_sheets = _get_excel_sheet_names(file_path)
            if sheet not in excel_sheets:
                print(f"[错误] Sheet '{sheet}' 在 Excel 文件中不存在")
                print(f"[Excel 中的 Sheet] {', '.join(excel_sheets)}")
                print(f"[已配置的 Sheet] {', '.join(sheets.keys())}")
            else:
                print(f"[状态] Sheet '{sheet}' 存在但尚未配置，以下是侦察结果：\n")
                scout_data = do_scout(file_path, sheet=sheet)

                # 自动推荐该 Sheet 的配置
                target_data = {k: v for k, v in scout_data.items() if k == sheet}
                recommended = _guess_config(target_data)
                sheet_cfg_rec = recommended["sheets"].get(sheet, {})
                sheet_json = json.dumps(sheet_cfg_rec, ensure_ascii=False, indent=2)

                abs_file = os.path.abspath(file_path)
                print(f"\n{'='*60}")
                print(f"[推荐配置] 工具已自动分析 Sheet '{sheet}'")
                print(f"[操作步骤]")
                print(f"  1. 读取现有配置文件：{config_path}")
                print(f'  2. 在 "sheets" 中添加以下内容：')
                print(f'     "{sheet}": {sheet_json}')
                print(f"  3. 用 Write 工具保存配置文件")
                print(f"\n[下一步] 保存后执行：")
                print(f'  python {TOOL_PATH} auto {abs_file} headers --sheet "{sheet}"')
                print(f"\n[已配置的 Sheet] {', '.join(sheets.keys())}")
            return
        target_sheets = {sheet: sheets[sheet]}
    else:
        if len(sheets) == 1:
            name = list(sheets.keys())[0]
            target_sheets = {name: sheets[name]}
            print(f"[提示] 只有一个 Sheet，已自动选择 \"{name}\"。多 Sheet 时必须用 --sheet 指定\n")
        else:
            if action == "query":
                print(f"[错误] 有多个 Sheet，查询时请用 --sheet 指定")
                print(f"[可用 Sheet] {', '.join(sheets.keys())}")
                return
            target_sheets = sheets

    print(f"[引擎] 读取={READ_ENGINE}, 处理=pandas")
    print(f"[配置] {config_path}")

    # 提示已有的 steps 文件
    steps_files = discover_steps_files(file_path)
    if steps_files:
        print(f"[可用规则] {', '.join(os.path.basename(f) for f in steps_files)}")
    print()

    for s_name, s_cfg in target_sheets.items():
        if len(target_sheets) > 1:
            print(f"\n{'='*40} Sheet: {s_name} {'='*40}")

        df = read_to_dataframe(file_path, s_name, s_cfg)

        if action == "headers":
            print(f"列名：{', '.join(df.columns.tolist())}")
            print(f"共 {len(df)} 行数据")
            print(f"\n列详情：")
            for col in df.columns:
                non_null = df[col].notna().sum()
                print(f"  {col}: {non_null}/{len(df)} 非空, 类型={df[col].dtype}")
            abs_file = os.path.abspath(file_path)
            sheet_opt = f' --sheet "{s_name}"'
            print(f"\n[下一步]")
            print(f"  预览数据: python {TOOL_PATH} auto {abs_file} preview{sheet_opt}")
            print(f"  条件查询: python {TOOL_PATH} auto {abs_file} query{sheet_opt} --where-col \"列名\" --where-op \">\" --where-val \"值\"")
            print(f"  清洗导出: python {TOOL_PATH} clean {abs_file}{sheet_opt} --preview")

        elif action == "preview":
            n = kwargs.get("n", 5)
            print(f"列名：{', '.join(df.columns.tolist())}")
            print(f"共 {len(df)} 行，预览前 {n} 行：\n")
            print(df.head(n).to_string(index=False))
            abs_file = os.path.abspath(file_path)
            sheet_opt = f' --sheet "{s_name}"'
            print(f"\n[下一步]")
            print(f"  条件查询: python {TOOL_PATH} auto {abs_file} query{sheet_opt} --where-col \"列名\" --where-op \">\" --where-val \"值\"")
            print(f"  清洗导出: python {TOOL_PATH} clean {abs_file}{sheet_opt} --preview")

        elif action == "query":
            _do_query(df, **kwargs)


def _do_query(df, **kwargs):
    where_col = kwargs.get("where_col")
    where_op = kwargs.get("where_op")
    where_val = kwargs.get("where_val")
    select_cols = kwargs.get("columns")
    sort_col = kwargs.get("sort")
    top = kwargs.get("top", 10)

    result = df

    # 筛选
    if where_col and where_op and where_val is not None:
        col_data = pd.to_numeric(result[where_col], errors="coerce")
        is_numeric = col_data.notna().any()

        if is_numeric and where_op in (">", "<", ">=", "<=", "==", "!="):
            val_num = float(where_val)
            ops = {">": "gt", "<": "lt", ">=": "ge", "<=": "le", "==": "eq", "!=": "ne"}
            result = result[getattr(col_data, ops[where_op])(val_num)]
        else:
            str_data = result[where_col].astype(str)
            if where_op == "==":
                result = result[str_data == where_val]
            elif where_op == "!=":
                result = result[str_data != where_val]
            elif where_op == "contains":
                result = result[str_data.str.contains(where_val, na=False)]
            elif where_op == "not_contains":
                result = result[~str_data.str.contains(where_val, na=False)]
            elif where_op == "startswith":
                result = result[str_data.str.startswith(where_val, na=False)]
            elif where_op == "endswith":
                result = result[str_data.str.endswith(where_val, na=False)]

    # 排序
    if sort_col:
        desc = sort_col.startswith("desc:")
        col = sort_col[5:] if desc else sort_col
        result = result.sort_values(col, ascending=not desc,
                                    key=lambda x: pd.to_numeric(x, errors="coerce"))

    # 选列
    if select_cols:
        cols = [c.strip() for c in select_cols.split(",")]
        result = result[cols]

    print(f"筛选后 {len(result)} 条，显示前 {top} 条：\n")
    print(result.head(top).to_string(index=False))


# ==================== clean 命令：pandas 清洗 ====================

def do_clean(file_path, rules_path=None, output_path=None, preview_only=False, sheet=None):
    config_path = get_config_path(file_path)
    if not os.path.exists(config_path):
        abs_file = os.path.abspath(file_path)
        print(f"[错误] 请先执行 auto 命令生成结构配置")
        print(f"  python {TOOL_PATH} auto {abs_file}")
        return

    # 确定清洗规则路径
    if rules_path is None:
        steps_files = discover_steps_files(file_path)
        if len(steps_files) == 0:
            prefix = get_steps_prefix(file_path)
            abs_file = os.path.abspath(file_path)
            sheet_opt = f' --sheet "{sheet}"' if sheet else ""
            print(f"[状态] 未找到清洗规则文件")
            print(f"\n{'='*60}")
            print(f"[命名规则] {{Excel文件名}}-{{操作描述}}.excel-steps.json")
            print(f"[示例文件名]")
            print(f"  {os.path.basename(prefix)}-按区域汇总.excel-steps.json")
            print(f"  {os.path.basename(prefix)}-数据清洗.excel-steps.json")
            print(f"\n[操作步骤]")
            print(f"  1. 根据数据特征，编写清洗规则 JSON（格式见下方）")
            print(f"  2. 用 Write 工具保存到 Excel 同目录")
            print(f"  3. 保存后执行：")
            print(f"     python {TOOL_PATH} clean {abs_file}{sheet_opt} --preview")
            print(f"\n[规则格式] steps 数组，按顺序执行：")
            print(f'{{"steps": [')
            print(f'  {{"action": "trim"}},')
            print(f'  {{"action": "replace", "column": "区域", "mapping": {{"旧值": "新值"}}}},')
            print(f'  {{"action": "fill_empty", "column": "销量", "value": 0}},')
            print(f'  {{"action": "dedup", "columns": ["区域", "型号"]}},')
            print(f'  {{"action": "filter", "conditions": [{{"column": "销量", "op": ">", "value": "0"}}], "logic": "and"}},')
            print(f'  {{"action": "sort", "column": "销量", "desc": true}},')
            print(f'  {{"action": "aggregate", "group_by": ["区域"], "metrics": {{"销量": "sum"}}}}')
            print(f"]}}")
            print(f"\n[更多操作] regex_replace, add_column, drop_columns, rename, type_convert, pivot")
            print(f"  查看详情: python {TOOL_PATH} help <操作名>")
            return
        elif len(steps_files) == 1:
            rules_path = steps_files[0]
            print(f"[自动发现] {os.path.basename(rules_path)}")
        else:
            abs_file = os.path.abspath(file_path)
            sheet_opt = f' --sheet "{sheet}"' if sheet else ""
            print(f"[发现多个规则文件] 请指定要使用的规则：")
            for f in steps_files:
                print(f"  python {TOOL_PATH} clean {abs_file} {f}{sheet_opt} --preview")
            return

    cfg = _load_config(config_path)
    sheets = cfg["sheets"]

    # 确定 sheet
    if sheet:
        if sheet not in sheets:
            excel_sheets = _get_excel_sheet_names(file_path)
            if sheet not in excel_sheets:
                print(f"[错误] Sheet '{sheet}' 在 Excel 文件中不存在")
                print(f"[Excel 中的 Sheet] {', '.join(excel_sheets)}")
                print(f"[已配置的 Sheet] {', '.join(sheets.keys())}")
            else:
                abs_file = os.path.abspath(file_path)
                print(f"[错误] Sheet '{sheet}' 尚未配置")
                print(f"[操作步骤] 先执行 auto 命令触发侦察并获取推荐配置：")
                print(f'  python {TOOL_PATH} auto {abs_file} headers --sheet "{sheet}"')
            return
        sheet_name = sheet
        sheet_cfg = sheets[sheet]
    else:
        if len(sheets) == 1:
            sheet_name = list(sheets.keys())[0]
            sheet_cfg = sheets[sheet_name]
        else:
            print(f"[错误] 有多个 Sheet，请用 --sheet 指定")
            print(f"[可用 Sheet] {', '.join(sheets.keys())}")
            return

    with open(rules_path, encoding="utf-8") as f:
        rules = json.load(f)

    # 校验 steps 格式
    errors = _validate_steps(rules)
    if errors:
        print(f"[规则校验失败] {rules_path}")
        for e in errors:
            print(f"  {e}")
        print(f"\n[提示] 用 help 查看正确格式：")
        # 提取出错的 action 名称用于提示
        actions_mentioned = set()
        for step in rules.get("steps", []):
            a = step.get("action")
            if a and a in _ACTION_REQUIRED:
                actions_mentioned.add(a)
        if actions_mentioned:
            for a in sorted(actions_mentioned):
                print(f"  python {TOOL_PATH} help {a}")
        else:
            print(f"  python {TOOL_PATH} help")
        return

    df = read_to_dataframe(file_path, sheet_name, sheet_cfg)
    print(f"[引擎] 读取={READ_ENGINE}, 处理=pandas")
    print(f"[规则] {rules_path}")
    print(f"[Sheet] {sheet_name}")
    print(f"[清洗前] {len(df)} 行 × {len(df.columns)} 列")

    for i, step in enumerate(rules.get("steps", [])):
        action = step.get("action")
        before = len(df)

        if action == "trim":
            cols = step.get("columns", df.select_dtypes(include=["object", "str"]).columns.tolist())
            for c in cols:
                if c in df.columns and c in df.select_dtypes(include=["object", "str"]).columns:
                    df[c] = df[c].astype(str).str.strip()
            print(f"  步骤{i+1} [去空格] {cols}")

        elif action == "replace":
            col = step["column"]
            mapping = step["mapping"]
            df[col] = df[col].astype(str).replace(mapping)
            print(f"  步骤{i+1} [替换] {col}: {len(mapping)}个映射规则")

        elif action == "fill_empty":
            col = step["column"]
            fill_val = step["value"]
            df[col] = df[col].replace(["", "None", "nan"], pd.NA)
            df[col] = df[col].fillna(fill_val)
            print(f"  步骤{i+1} [填充空值] {col}: 填充为 {fill_val}")

        elif action == "dedup":
            dedup_cols = step.get("columns")
            before_len = len(df)
            df = df.drop_duplicates(subset=dedup_cols, keep="first")
            print(f"  步骤{i+1} [去重] 按{dedup_cols or '全列'}: 移除{before_len - len(df)}条")

        elif action == "filter":
            conditions = step.get("conditions", [])
            logic = step.get("logic", "and")
            masks = []
            for cond in conditions:
                col, op, val = cond["column"], cond["op"], cond["value"]
                col_num = pd.to_numeric(df[col], errors="coerce")
                is_numeric = col_num.notna().any()

                if is_numeric and op in (">", "<", ">=", "<=", "==", "!="):
                    val_n = float(val)
                    op_map = {">": "gt", "<": "lt", ">=": "ge", "<=": "le", "==": "eq", "!=": "ne"}
                    masks.append(getattr(col_num, op_map[op])(val_n))
                else:
                    s = df[col].astype(str)
                    if op == "==":           masks.append(s == val)
                    elif op == "!=":         masks.append(s != val)
                    elif op == "contains":   masks.append(s.str.contains(val, na=False))
                    elif op == "not_contains": masks.append(~s.str.contains(val, na=False))
                    elif op == "startswith": masks.append(s.str.startswith(val, na=False))
                    elif op == "endswith":   masks.append(s.str.endswith(val, na=False))

            if masks:
                combined = masks[0]
                for m in masks[1:]:
                    combined = (combined & m) if logic == "and" else (combined | m)
                df = df[combined]
            print(f"  步骤{i+1} [筛选] {logic.upper()} {len(conditions)}条件: {before}→{len(df)}行")

        elif action == "regex_replace":
            col = step["column"]
            df[col] = df[col].astype(str).str.replace(
                step["pattern"], step["replacement"], regex=True
            )
            print(f"  步骤{i+1} [正则替换] {col}")

        elif action == "add_column":
            col_name = step["name"]
            formula = step["formula"]
            rnd = step.get("round")
            # 构建 pandas eval 表达式：{列名} → `列名`
            expr = formula
            for h in df.columns:
                expr = expr.replace("{" + h + "}", f"`{h}`")
            try:
                df[col_name] = df.eval(expr)
                if rnd is not None:
                    df[col_name] = df[col_name].round(rnd)
            except Exception as e:
                df[col_name] = pd.NA
                print(f"    警告: 公式计算失败 - {e}")
            print(f"  步骤{i+1} [新增列] {col_name}")

        elif action == "drop_columns":
            drop = step["columns"]
            df = df.drop(columns=[c for c in drop if c in df.columns])
            print(f"  步骤{i+1} [删列] {drop}")

        elif action == "sort":
            col = step["column"]
            desc = step.get("desc", False)
            df = df.sort_values(col, ascending=not desc,
                                key=lambda x: pd.to_numeric(x, errors="coerce"))
            print(f"  步骤{i+1} [排序] {col} {'降序' if desc else '升序'}")

        elif action == "aggregate":
            group_by = step["group_by"]
            metrics = step["metrics"]
            # 先转数值列
            for col in metrics:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            agg_map = {}
            for col, func in metrics.items():
                if func == "avg":
                    agg_map[col] = "mean"
                else:
                    agg_map[col] = func
            df = df.groupby(group_by, as_index=False).agg(agg_map)
            print(f"  步骤{i+1} [聚合] 按{group_by}: {len(df)}组")

        elif action == "rename":
            df = df.rename(columns=step["mapping"])
            print(f"  步骤{i+1} [重命名] {step['mapping']}")

        elif action == "type_convert":
            for col, dtype in step["columns"].items():
                if dtype in ("int", "float"):
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                    if dtype == "int":
                        df[col] = df[col].fillna(0).astype(int)
                elif dtype == "datetime":
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                elif dtype == "str":
                    df[col] = df[col].astype(str)
            print(f"  步骤{i+1} [类型转换] {step['columns']}")

        elif action == "pivot":
            df = pd.pivot_table(df,
                                index=step["index"],
                                columns=step["columns"],
                                values=step["values"],
                                aggfunc=step.get("aggfunc", "sum")).reset_index()
            df.columns = [str(c) if not isinstance(c, tuple) else "_".join(str(x) for x in c)
                          for c in df.columns]
            print(f"  步骤{i+1} [透视] {step['index']} × {step['columns']}")

        else:
            valid = "trim replace fill_empty dedup filter regex_replace add_column drop_columns sort aggregate rename type_convert pivot"
            print(f"  步骤{i+1} [错误] 未知操作 '{action}'，跳过")
            print(f"         [可用操作] {valid}")

    # 四舍五入浮点显示
    float_cols = df.select_dtypes(include="float").columns
    df[float_cols] = df[float_cols].round(2)

    print(f"\n[清洗后] {len(df)} 行 × {len(df.columns)} 列")

    # 预览
    if preview_only or not output_path:
        print(f"\n预览前 10 行：")
        print(df.head(10).to_string(index=False))
        if not output_path:
            print(f"\n[提示] 未指定输出路径，仅预览。用 -o 指定输出文件。")
        return

    # 导出
    ext = os.path.splitext(output_path)[1].lower()
    if ext == ".csv":
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
    elif ext == ".json":
        df.to_json(output_path, orient="records", force_ascii=False, indent=2)
    elif ext in (".xlsx", ".xls"):
        df.to_excel(output_path, index=False, engine="openpyxl")
    else:
        print(f"[错误] 不支持的格式: {ext}，支持 .csv .json .xlsx")
        return
    print(f"[导出] {output_path} ({len(df)}行)")


# ==================== export 命令：直接导出干净数据（供自定义脚本使用）====================

def do_export(file_path, output_path, sheet=None):
    config_path = get_config_path(file_path)
    if not os.path.exists(config_path):
        abs_file = os.path.abspath(file_path)
        print(f"[错误] 请先执行 auto 命令生成结构配置")
        print(f"  python {TOOL_PATH} auto {abs_file}")
        return

    cfg = _load_config(config_path)
    sheets = cfg["sheets"]

    if sheet:
        if sheet not in sheets:
            print(f"[错误] Sheet '{sheet}' 未配置")
            print(f"[已配置的 Sheet] {', '.join(sheets.keys())}")
            return
        sheet_name = sheet
        sheet_cfg = sheets[sheet]
    else:
        if len(sheets) == 1:
            sheet_name = list(sheets.keys())[0]
            sheet_cfg = sheets[sheet_name]
        else:
            print(f"[错误] 有多个 Sheet，请用 --sheet 指定")
            print(f"[可用 Sheet] {', '.join(sheets.keys())}")
            return

    df = read_to_dataframe(file_path, sheet_name, sheet_cfg)

    # 四舍五入浮点显示
    float_cols = df.select_dtypes(include="float").columns
    df[float_cols] = df[float_cols].round(2)

    print(f"[引擎] 读取={READ_ENGINE}, 处理=pandas")
    print(f"[Sheet] {sheet_name}")
    print(f"[数据] {len(df)} 行 × {len(df.columns)} 列")

    ext = os.path.splitext(output_path)[1].lower()
    if ext == ".csv":
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
    elif ext == ".json":
        df.to_json(output_path, orient="records", force_ascii=False, indent=2)
    elif ext in (".xlsx", ".xls"):
        df.to_excel(output_path, index=False, engine="openpyxl")
    else:
        print(f"[错误] 不支持的格式: {ext}，支持 .csv .json .xlsx")
        return
    print(f"[导出] {output_path} ({len(df)}行)")


# ==================== steps 校验 ====================

_ACTION_REQUIRED = {
    "trim": [],
    "replace": ["column", "mapping"],
    "fill_empty": ["column", "value"],
    "dedup": [],
    "filter": ["conditions"],
    "regex_replace": ["column", "pattern", "replacement"],
    "add_column": ["name", "formula"],
    "drop_columns": ["columns"],
    "sort": ["column"],
    "aggregate": ["group_by", "metrics"],
    "rename": ["mapping"],
    "type_convert": ["columns"],
    "pivot": ["index", "columns", "values"],
}


def _validate_steps(rules):
    """校验 steps 格式，返回错误列表。空列表表示通过。"""
    errors = []
    steps = rules.get("steps")
    if steps is None:
        errors.append("缺少 \"steps\" 字段（应为数组）")
        return errors
    if not isinstance(steps, list):
        errors.append("\"steps\" 应为数组")
        return errors

    for i, step in enumerate(steps, 1):
        action = step.get("action")
        if not action:
            errors.append(f"步骤{i}: 缺少 action 字段")
            continue
        if action not in _ACTION_REQUIRED:
            errors.append(
                f"步骤{i}: 未知操作 '{action}'，"
                f"可用: {', '.join(_ACTION_REQUIRED.keys())}"
            )
            continue
        for param in _ACTION_REQUIRED[action]:
            if param not in step:
                errors.append(f"步骤{i} [{action}]: 缺少必填参数 '{param}'")

    return errors


# ==================== help 命令 ====================

_HELP_ACTIONS = {
    "trim": """[trim - 去首尾空格]

格式:
  {"action": "trim"}
  {"action": "trim", "columns": ["区域", "型号"]}

不指定 columns 则处理所有字符串列。""",

    "replace": """[replace - 批量精确替换]

格式:
  {"action": "replace", "column": "区域", "mapping": {"华东区": "华东", "华南地区": "华南"}}

参数:
  column:  目标列名
  mapping: {"旧值": "新值"} 字典""",

    "fill_empty": """[fill_empty - 填充空值]

格式:
  {"action": "fill_empty", "column": "销量", "value": 0}

空字符串、None、NaN 都会被填充。
参数:
  column: 目标列名
  value:  填充值（数字或字符串）""",

    "dedup": """[dedup - 去重]

格式:
  {"action": "dedup"}
  {"action": "dedup", "columns": ["区域", "型号"]}

不指定 columns 则按全列去重，保留第一条。""",

    "filter": """[filter - 多条件筛选]

格式:
  {"action": "filter", "conditions": [...], "logic": "and"}

conditions 中每个条件:
  {"column": "列名", "op": "运算符", "value": "值"}

运算符: > < >= <= == != contains not_contains startswith endswith
logic:  and（全部满足）或 or（任一满足）
数值列自动用数值比较。

示例:
  {"action": "filter", "conditions": [
    {"column": "销量", "op": ">", "value": "1000"},
    {"column": "类别", "op": "==", "value": "笔记本"}
  ], "logic": "and"}""",

    "regex_replace": r"""[regex_replace - 正则替换]

格式:
  {"action": "regex_replace", "column": "电话", "pattern": "^(\\d{3})\\d{4}(\\d{4})$", "replacement": "\\1****\\2"}

参数:
  column:      目标列名
  pattern:     正则表达式
  replacement: 替换字符串（支持 \\1 反向引用）""",

    "add_column": """[add_column - 新增计算列]

格式:
  {"action": "add_column", "name": "利润率", "formula": "({营收}-{成本})/{营收}*100", "round": 2}

参数:
  name:    新列名
  formula: 计算公式，用 {列名} 引用已有列，支持 + - * / **
  round:   (可选) 小数位数""",

    "drop_columns": """[drop_columns - 删除列]

格式:
  {"action": "drop_columns", "columns": ["序号", "备注"]}""",

    "sort": """[sort - 排序]

格式:
  {"action": "sort", "column": "销量", "desc": true}

参数:
  column: 排序列
  desc:   true 降序，false 或省略为升序
自动转数值排序，非数值排末尾。""",

    "aggregate": """[aggregate - 分组聚合]

格式:
  {"action": "aggregate", "group_by": ["区域"], "metrics": {"销量": "sum", "单价": "avg"}}

参数:
  group_by: 分组列（数组）
  metrics:  {"列名": "聚合函数"} 字典
聚合函数: sum avg count max min""",

    "rename": """[rename - 重命名列]

格式:
  {"action": "rename", "mapping": {"旧列名": "新列名"}}""",

    "type_convert": """[type_convert - 类型转换]

格式:
  {"action": "type_convert", "columns": {"销量": "int", "日期": "datetime"}}

支持类型: int float datetime str
转换失败的值变为 NaN/NaT。""",

    "pivot": """[pivot - 数据透视]

格式:
  {"action": "pivot", "index": "区域", "columns": "类别", "values": "销量", "aggfunc": "sum"}

参数:
  index:   行索引列
  columns: 列头列
  values:  值列
  aggfunc: 聚合函数（sum mean count max min）""",
}


def do_help(topic=None):
    if topic is None:
        print("[可用清洗操作]")
        descs = {
            "trim": "去首尾空格", "replace": "批量精确替换", "fill_empty": "填充空值",
            "dedup": "去重", "filter": "多条件筛选", "regex_replace": "正则替换",
            "add_column": "新增计算列", "drop_columns": "删除列", "sort": "排序",
            "aggregate": "分组聚合", "rename": "重命名列", "type_convert": "类型转换",
            "pivot": "数据透视",
        }
        for name, desc in descs.items():
            print(f"  {name:<16}{desc}")
        print(f"\n[用法]")
        print(f"  python {TOOL_PATH} help <操作名>          ← 查看具体格式")
        print(f"  python {TOOL_PATH} help custom-scripts    ← 自定义脚本指南")
    elif topic == "custom-scripts":
        print("[自定义脚本指南]")
        print()
        print("当内置 13 种操作无法满足需求时（多表关联、复杂分支、统计可视化等），")
        print("用自定义 Python 脚本处理。")
        print()
        print("工作链路:")
        print("  excel_tool.py export → 干净 csv/json → 自定义脚本处理 → 输出结果")
        print()
        print("步骤:")
        print(f"  1. 导出干净数据:")
        print(f"     python {TOOL_PATH} export <Excel文件> -o data.csv --sheet \"Sheet名\"")
        print(f"  2. 在 Excel 同目录下创建 scripts/ 文件夹")
        print(f"  3. 编写 Python 脚本（读取 csv/json，处理，输出结果）:")
        print()
        print("     import pandas as pd")
        print("     df = pd.read_csv(\"data.csv\")")
        print("     # 处理逻辑...")
        print("     result.to_csv(\"output.csv\", index=False, encoding=\"utf-8-sig\")")
        print()
        print("约定:")
        print("  - 脚本读取 export 导出的干净 csv/json，不直接读 .xlsx")
        print("  - 文件名要描述功能（如 跨表关联分析.py），不要用 temp.py")
        print("  - 禁止修改原始 Excel 文件")
        print("  - 可使用 pandas、numpy、openpyxl 等已安装的包")
    elif topic in _HELP_ACTIONS:
        print(_HELP_ACTIONS[topic])
    else:
        print(f"[错误] 未知主题 '{topic}'")
        print(f"\n可用操作: {', '.join(_HELP_ACTIONS.keys())}")
        print(f"特殊主题: custom-scripts")
        print(f"\n用法: python {TOOL_PATH} help <操作名>")


# ==================== CLI ====================

def main():
    parser = argparse.ArgumentParser(description="Excel 报表工具")
    sub = parser.add_subparsers(dest="command")

    p_scout = sub.add_parser("scout", help="侦察：原样读取前N行")
    p_scout.add_argument("file")
    p_scout.add_argument("-n", type=int, default=8)
    p_scout.add_argument("--sheet", help="指定 Sheet 名称")

    p_auto = sub.add_parser("auto", help="自动模式")
    p_auto.add_argument("file")
    p_auto.add_argument("action", nargs="?", default="preview",
                        choices=["headers", "preview", "query"])
    p_auto.add_argument("-n", type=int, default=5)
    p_auto.add_argument("--sheet", help="指定 Sheet 名称")
    p_auto.add_argument("--where-col")
    p_auto.add_argument("--where-op")
    p_auto.add_argument("--where-val")
    p_auto.add_argument("-c", "--columns")
    p_auto.add_argument("-s", "--sort", help="排序，降序用 desc:列名")
    p_auto.add_argument("-t", "--top", type=int, default=10)

    p_clean = sub.add_parser("clean", help="按规则清洗数据")
    p_clean.add_argument("file")
    p_clean.add_argument("rules", nargs="?", default=None,
                         help="规则文件路径（可选，省略时自动查找 .excel-steps.json）")
    p_clean.add_argument("-o", "--output")
    p_clean.add_argument("--preview", action="store_true")
    p_clean.add_argument("--sheet", help="指定 Sheet 名称")

    p_export = sub.add_parser("export", help="导出干净数据（供自定义脚本使用）")
    p_export.add_argument("file")
    p_export.add_argument("-o", "--output", required=True, help="输出路径（.csv/.json/.xlsx）")
    p_export.add_argument("--sheet", help="指定 Sheet 名称")

    p_cfg = sub.add_parser("config-path")
    p_cfg.add_argument("file")

    p_steps = sub.add_parser("steps-path", help="查看清洗规则文件的命名模式和已有文件")
    p_steps.add_argument("file")

    p_help = sub.add_parser("help", help="查看操作格式")
    p_help.add_argument("topic", nargs="?", default=None,
                        help="操作名（如 filter）或 custom-scripts")

    args = parser.parse_args()

    if args.command == "scout":
        print(f"[引擎] {READ_ENGINE}")
        do_scout(args.file, args.n, sheet=args.sheet)
    elif args.command == "auto":
        do_auto(args.file, args.action, sheet=args.sheet, n=args.n,
                where_col=args.where_col, where_op=args.where_op,
                where_val=args.where_val, columns=args.columns,
                sort=args.sort, top=args.top)
    elif args.command == "clean":
        do_clean(args.file, args.rules, args.output, args.preview, sheet=args.sheet)
    elif args.command == "export":
        do_export(args.file, args.output, sheet=args.sheet)
    elif args.command == "config-path":
        print(get_config_path(args.file))
    elif args.command == "steps-path":
        prefix = get_steps_prefix(args.file)
        print(f"[命名规则] {os.path.basename(prefix)}-<操作描述>.excel-steps.json")
        steps_files = discover_steps_files(args.file)
        if steps_files:
            print(f"[已有文件]")
            for f in steps_files:
                print(f"  - {os.path.basename(f)}")
        else:
            print(f"[已有文件] 无")
    elif args.command == "help":
        do_help(args.topic)
    else:
        parser.print_help()


if __name__ == "__main__":
    try:
        main()
    except FileNotFoundError as e:
        print(f"[错误] 文件不存在: {e.filename}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"[错误] JSON 解析失败: {e.msg}（行{e.lineno} 列{e.colno}）")
        print(f"[提示] 检查 JSON 文件语法，常见问题：多余逗号、缺少引号、中文引号")
        sys.exit(1)
    except KeyError as e:
        print(f"[错误] 列名不存在: {e}")
        print(f"[提示] 用 auto headers 命令查看可用列名")
        sys.exit(1)
    except Exception as e:
        print(f"[错误] {type(e).__name__}: {e}")
        sys.exit(1)
